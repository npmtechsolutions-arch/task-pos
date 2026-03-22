"""
Export service — multi-format report export.

Supported: CSV, Excel (openpyxl), JSON, PDF (reportlab with HTML fallback)
"""

import csv
import io
import json
from datetime import datetime
from typing import Any, Dict, List, Optional

from app.core.logging import get_logger

logger = get_logger(__name__)


class ExportService:
    """Generate downloadable report files in multiple formats."""

    # ── CSV ────────────────────────────────────────────────────────────────────

    def to_csv(self, rows: List[Dict[str, Any]], title: str = "report") -> bytes:
        """Return UTF-8 CSV bytes."""
        if not rows:
            return b""
        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=list(rows[0].keys()), extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)
        return output.getvalue().encode("utf-8")

    # ── JSON ───────────────────────────────────────────────────────────────────

    def to_json(self, rows: List[Dict[str, Any]], title: str = "report") -> bytes:
        """Return pretty-printed JSON bytes."""
        return json.dumps(
            {"title": title, "generated_at": datetime.utcnow().isoformat(), "rows": rows},
            indent=2,
            default=str,
        ).encode("utf-8")

    # ── Excel ──────────────────────────────────────────────────────────────────

    def to_excel(self, rows: List[Dict[str, Any]], title: str = "Report") -> bytes:
        """Return .xlsx bytes using openpyxl."""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter
        except ImportError:
            logger.warning("openpyxl not installed — falling back to CSV for Excel export")
            return self.to_csv(rows, title)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = title[:31]  # Excel sheet name max 31 chars

        if not rows:
            buf = io.BytesIO()
            wb.save(buf)
            return buf.getvalue()

        headers = list(rows[0].keys())

        # Header row styling
        HEADER_FILL = PatternFill("solid", fgColor="4F46E5")
        HEADER_FONT = Font(bold=True, color="FFFFFF")

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=str(header).replace("_", " ").title())
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center")

        # Data rows
        for row_idx, row in enumerate(rows, 2):
            for col_idx, key in enumerate(headers, 1):
                value = row.get(key)
                if isinstance(value, datetime):
                    value = value.isoformat()
                ws.cell(row=row_idx, column=col_idx, value=value)

        # Auto-width
        for col_idx in range(1, len(headers) + 1):
            max_len = max(
                len(str(ws.cell(row=r, column=col_idx).value or ""))
                for r in range(1, min(len(rows) + 2, 50))
            )
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 40)

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    # ── PDF ────────────────────────────────────────────────────────────────────

    def to_pdf(self, rows: List[Dict[str, Any]], title: str = "Report") -> bytes:
        """Return PDF bytes using reportlab (falls back to HTML if unavailable)."""
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.lib.styles import getSampleStyleSheet
            from reportlab.lib.units import cm
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        except ImportError:
            logger.warning("reportlab not installed — generating HTML PDF substitute")
            return self._html_fallback(rows, title)

        buf = io.BytesIO()
        doc = SimpleDocTemplate(
            buf,
            pagesize=landscape(A4),
            title=title,
            topMargin=1.5 * cm,
            bottomMargin=1.5 * cm,
            leftMargin=2 * cm,
            rightMargin=2 * cm,
        )

        styles = getSampleStyleSheet()
        story = [
            Paragraph(title, styles["Title"]),
            Paragraph(
                f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}",
                styles["Normal"]
            ),
            Spacer(1, 0.5 * cm),
        ]

        if rows:
            headers = list(rows[0].keys())
            table_data = [[h.replace("_", " ").title() for h in headers]]
            for row in rows[:500]:  # cap at 500 rows for PDF
                table_data.append([str(row.get(h, "")) for h in headers])

            col_count = len(headers)
            col_width = (landscape(A4)[0] - 4 * cm) / col_count

            t = Table(table_data, colWidths=[col_width] * col_count, repeatRows=1)
            t.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4F46E5")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 9),
                ("FONTSIZE", (0, 1), (-1, -1), 8),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F3F4F6")]),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#E5E7EB")),
                ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ]))
            story.append(t)

        doc.build(story)
        return buf.getvalue()

    def _html_fallback(self, rows: List[Dict[str, Any]], title: str) -> bytes:
        """Return simple HTML as fallback when reportlab unavailable."""
        headers = list(rows[0].keys()) if rows else []
        header_html = "".join(f"<th>{h}</th>" for h in headers)
        rows_html = ""
        for row in rows[:500]:
            cells = "".join(f"<td>{row.get(h, '')}</td>" for h in headers)
            rows_html += f"<tr>{cells}</tr>"
        html = f"""<!DOCTYPE html><html><head><title>{title}</title>
<style>body{{font-family:sans-serif}}table{{border-collapse:collapse;width:100%}}
th{{background:#4F46E5;color:white;padding:8px}}td{{padding:6px;border:1px solid #ddd}}</style>
</head><body><h1>{title}</h1><p>Generated: {datetime.utcnow().isoformat()}</p>
<table><thead><tr>{header_html}</tr></thead><tbody>{rows_html}</tbody></table></body></html>"""
        return html.encode("utf-8")

    def export(self, rows: List[Dict[str, Any]], fmt: str, title: str = "Report") -> tuple[bytes, str, str]:
        """
        Returns (file_bytes, content_type, filename).
        fmt: csv | excel | json | pdf
        """
        ts = datetime.utcnow().strftime("%Y%m%d_%H%M")
        safe_title = "".join(c if c.isalnum() else "_" for c in title)[:40]

        if fmt == "csv":
            return self.to_csv(rows, title), "text/csv", f"{safe_title}_{ts}.csv"
        elif fmt == "excel":
            return self.to_excel(rows, title), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", f"{safe_title}_{ts}.xlsx"
        elif fmt == "json":
            return self.to_json(rows, title), "application/json", f"{safe_title}_{ts}.json"
        elif fmt == "pdf":
            return self.to_pdf(rows, title), "application/pdf", f"{safe_title}_{ts}.pdf"
        else:
            raise ValueError(f"Unknown format: {fmt}. Use csv | excel | json | pdf")
